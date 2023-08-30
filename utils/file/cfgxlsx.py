#  读取其他数据资源
import os

from openpyxl import load_workbook

from common import const
from module import tableCfg
from utils.file.base import filePath2FileNameExist


def getCfgPath():
    return const.CFG_DIR_PATH


def readCfgData(file_path):
    # print(file_path)
    wb = load_workbook(filename=file_path)
    sheet = wb['Sheet1']

    field_list = []
    # 按列去读
    columns = sheet.iter_cols(min_col=1, max_col=4, values_only=True)
    for column in columns:
        #  1列4个关键字
        if len(column) < 4:
            print(f"error file_path:{file_path} column not equal 4")
            exit(-1)

        field_english_name = column[0]
        if field_english_name is None:
            continue

        field_type = column[2]
        # todo 后续可以扩展
        flag = False
        if field_type in const.EXPORT_GO_CODE_TYPE:
            flag = True
            field_type = const.EXPORT_GO_CODE_TYPE[field_type]

        if not flag:
            print(f"field type error file_path:{file_path} "
                  f"field_english_name:{field_english_name} "
                  f"field_type:{field_type}")
            exit(-1)

        obj = tableCfg.TableCfg(field_english_name, column[1], field_type, column[3])
        field_list.append(obj)
        # print(obj)

    file_name = filePath2FileNameExist(file_path)
    const.TABLE_DEFINE_MAP[file_name] = field_list

    table_data_values = []
    # 按行去读
    for row in sheet.iter_rows(min_row=5, values_only=True):
        data = []
        for i, cell in enumerate(row):
            col_note = sheet.cell(row=1, column=i + 1).value
            # 注释列跳过
            if col_note is None:
                continue
            data.append(cell)

        # print(data)
        table_data_values.append(data)


    const.TABLE_DATA_MAP[file_name] = table_data_values
    wb.close()


def readAllCfgData(path=getCfgPath()):
    # print("path:", path)

    for filename in os.listdir(path):
        # 如果当前路径表示一个目录，则递归读取该目录下的文件
        full_path = os.path.join(path, filename)
        if os.path.isdir(full_path):
            readAllCfgData(full_path)
            return

        # 是否是文件
        if os.path.isfile(full_path) is False:
            continue
        # 跳过临时文件
        if filename.startswith("~$"):
            print(f"waring {filename} reading")
            continue
        # 元数据跳过
        if filename.find(const.META_NAME) != -1:
            continue

        if filename.endswith(".xlsx"):
            readCfgData(full_path)
