# 先定义名字,中文表对应导出英文名称
import os.path

from common import const
from openpyxl import load_workbook


def mkdir(dirName):
    if not os.path.exists(dirName):
        os.makedirs(dirName)


def getMetaXlsxPath():
    return os.path.join(const.CFG_DIR_PATH, const.META_NAME)


def filePath2FileName(file_path):
    return os.path.basename(file_path).replace(".xlsx", "")


def filePath2FileNameExist(file_path):
    file_name = filePath2FileName(file_path)
    if file_name is None or file_name == "":
        print(f"error file_path:{file_path} not exist meta data")
        exit(-1)
        return None

    return file_name


def getExportNameByFileName(file_name):
    return const.META_BASE_MAP[file_name]


def getExportNameByFileNameExist(file_name):
    export_name = getExportNameByFileName(file_name)
    if export_name is None or export_name == "":
        print(f"error file_name{file_name} export name not exist")
        exit(-1)
        return None

    return export_name


def getExportNameByFilePath(file_path):
    file_name = filePath2FileName(file_path)
    if file_name is None or file_name == "":
        return None
    export_name = getExportNameByFileName(file_name)
    return export_name


def getExportNameByFilePathExist(file_path):
    file_name = filePath2FileNameExist(file_path)
    export_name = getExportNameByFileNameExist(file_name)

    return export_name


def readMetaXlsx():
    wb = load_workbook(filename=getMetaXlsxPath())
    sheet = wb['Sheet1']

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if len(row) != 2:
            print("error:meta data row not equal 2", row)
            exit(-1)

        key = row[0]
        value = row[1]

        if key is None:
            continue

        if const.META_BASE_MAP.get(key):
            print(f"error:key:{key} is repeat")
            exit(-1)

        if value is None:
            print(f"error: key:{key},but value is none")
            exit(-1)

        const.META_BASE_MAP[key] = value

    wb.close()
    # print(const.META_BASE_MAP)
